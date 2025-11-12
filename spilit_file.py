import argparse
import csv
import os
import re
from typing import Optional

import pandas as pd

try:  # pragma: no cover - dipende dall'ambiente dell'utente
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - openpyxl potrebbe non essere installato
    load_workbook = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
SUPPORTED_EXTS = {".xlsx", ".xlsm", ".xls", ".csv"}
EXCEL_WITH_FORMATS = {".xlsx", ".xlsm"}
HEADER_MODE_REPEAT = "repeat"
HEADER_MODE_FIRST_ONLY = "first-only"
HEADER_MODE_NONE = "none"
HEADER_MODE_FORMATTED = "formatted"
HEADER_MODE_FORMATTED_FIRST_ONLY = "formatted-first-only"
HEADER_MODE_CHOICES = (
    HEADER_MODE_REPEAT,
    HEADER_MODE_FIRST_ONLY,
    HEADER_MODE_NONE,
    HEADER_MODE_FORMATTED,
    HEADER_MODE_FORMATTED_FIRST_ONLY,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Divide un file in blocchi più piccoli da 100 righe (default)."
    )
    parser.add_argument(
        "--file",
        help=(
            "Percorso del file da suddividere. Se non specificato viene preso il "
            "primo file supportato presente nella cartella 'input'."
        ),
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=100,
        help="Numero di righe per ciascun file generato (default: 100).",
    )
    parser.add_argument(
        "--header-mode",
        choices=HEADER_MODE_CHOICES,
        default=HEADER_MODE_REPEAT,
        help=(
            "Gestione dell'intestazione nei file generati: "
            "'repeat' per ripeterla ovunque; "
            "'first-only' per mantenerla solo nel primo file; "
            "'none' per ometterla sempre; "
            "'formatted' per ripeterla ovunque dopo aver rimosso spazi e caratteri speciali; "
            "'formatted-first-only' per usare l'intestazione ripulita solo nel primo file."
        ),
    )
    return parser.parse_args()


def _select_file_from_input() -> str:
    if not os.path.isdir(INPUT_DIR):
        raise SystemExit(f"Errore: la cartella di input non esiste: {INPUT_DIR}")

    for entry in sorted(os.listdir(INPUT_DIR)):
        if os.path.splitext(entry)[1].lower() in SUPPORTED_EXTS:
            return os.path.join(INPUT_DIR, entry)

    raise SystemExit(
        "Errore: nessun file supportato trovato nella cartella 'input'. "
        f"Formati accettati: {', '.join(sorted(SUPPORTED_EXTS))}"
    )


def _resolve_file(path: Optional[str]) -> str:
    if path:
        absolute = os.path.abspath(path)
        if not os.path.isfile(absolute):
            raise SystemExit(f"Errore: il file '{path}' non esiste.")
        if os.path.splitext(absolute)[1].lower() not in SUPPORTED_EXTS:
            raise SystemExit(
                f"Errore: il file '{path}' non ha un'estensione supportata "
                f"({', '.join(sorted(SUPPORTED_EXTS))})."
            )
        return absolute
    return _select_file_from_input()


def _detect_csv_delimiter(file_path: str) -> str:
    """
    Prova a rilevare il delimitatore principale del file CSV
    restituendo una virgola come fallback.
    """
    try:
        with open(file_path, "r", newline="", encoding="utf-8-sig") as handle:
            sample = handle.read(4096)
            if not sample:
                return ","
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample, delimiters=[",", ";", "\t", "|", ":"])
            return dialect.delimiter
    except (OSError, csv.Error):
        return ","


def _load_dataframe(file_path: str) -> tuple[pd.DataFrame, str | None]:
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".csv":
            delimiter = _detect_csv_delimiter(file_path)
            df = pd.read_csv(file_path, sep=delimiter, engine="python")
            return df, delimiter
        df = pd.read_excel(file_path, header=0)
        return df, None
    except Exception as exc:  # pragma: no cover - dipende dai file utente
        raise SystemExit(f"Errore durante la lettura di '{file_path}': {exc}") from exc


def _format_headers(columns: pd.Index | list[str]) -> list[str]:
    formatted: list[str] = []
    used: set[str] = set()
    counts: dict[str, int] = {}

    for position, column in enumerate(columns, start=1):
        text = re.sub(r"[^0-9A-Za-z]", "", str(column))
        if not text:
            text = f"Colonna{position}"
        base = text
        while text in used:
            counts[base] = counts.get(base, 0) + 1
            text = f"{base}{counts[base]}"
        used.add(text)
        formatted.append(text)
    return formatted


def _extract_excel_column_formats(file_path: str) -> list[str]:
    """Legge i formati numerici delle colonne di un file Excel."""
    if load_workbook is None:
        return []
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=False)
    except Exception:  # pragma: no cover - dipende dai file utente
        return []

    sheet = workbook.active
    column_formats: list[str] = []
    for column in range(1, sheet.max_column + 1):
        detected_format: str | None = None
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column)
            if cell.value is None:
                continue
            fmt = cell.number_format or ""
            if fmt.lower() != "general":
                detected_format = fmt
                break
        column_formats.append(detected_format or "")

    workbook.close()
    return column_formats


def _apply_excel_column_formats(
    file_path: str, column_formats: list[str], header_present: bool
) -> None:
    if load_workbook is None or not column_formats:
        return
    try:
        workbook = load_workbook(file_path)
    except Exception:  # pragma: no cover - dipende dai file utente
        return

    sheet = workbook.active
    start_row = 2 if header_present else 1
    max_row = sheet.max_row
    for idx, fmt in enumerate(column_formats, start=1):
        if not fmt:
            continue
        for row in range(start_row, max_row + 1):
            sheet.cell(row=row, column=idx).number_format = fmt

    workbook.save(file_path)
    workbook.close()


def main() -> None:
    args = parse_args()
    file_path = _resolve_file(args.file)

    if args.chunk_size <= 0:
        raise SystemExit("Errore: --chunk-size deve essere un numero positivo.")

    df, delimiter = _load_dataframe(file_path)
    if df.empty:
        raise SystemExit("Errore: il file selezionato è vuoto.")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    base_name = os.path.basename(file_path)
    file_ext = os.path.splitext(base_name)[1].lower()
    chunk_size = args.chunk_size
    column_formats: list[str] = []
    if file_ext in EXCEL_WITH_FORMATS:
        column_formats = _extract_excel_column_formats(file_path)

    total_chunks = (len(df) + chunk_size - 1) // chunk_size
    format_header = args.header_mode in {
        HEADER_MODE_FORMATTED,
        HEADER_MODE_FORMATTED_FIRST_ONLY,
    }
    formatted_columns = _format_headers(df.columns) if format_header else None

    for index, start in enumerate(range(0, len(df), chunk_size), start=1):
        chunk = df.iloc[start : start + chunk_size].copy()
        if format_header and formatted_columns is not None:
            chunk.columns = formatted_columns
        output_file = os.path.join(OUTPUT_DIR, f"{index} di {total_chunks} {base_name}")
        write_header = (
            args.header_mode in (HEADER_MODE_REPEAT, HEADER_MODE_FORMATTED)
            or (
                args.header_mode
                in (HEADER_MODE_FIRST_ONLY, HEADER_MODE_FORMATTED_FIRST_ONLY)
                and index == 1
            )
        )
        try:
            if file_ext == ".csv":
                chunk.to_csv(
                    output_file,
                    index=False,
                    header=write_header,
                    sep=delimiter or ",",
                )
            else:
                chunk.to_excel(output_file, index=False, header=write_header)
                if column_formats:
                    _apply_excel_column_formats(
                        output_file, column_formats, header_present=write_header
                    )
        except Exception as exc:  # pragma: no cover - dipende dai file utente
            raise SystemExit(f"Errore durante il salvataggio di '{output_file}': {exc}") from exc
        print(f"Creato: {output_file}")


if __name__ == "__main__":
    main()
